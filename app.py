import os
import json
import time
import threading
import hmac
from pathlib import Path
from datetime import datetime, timedelta
from io import BytesIO

from flask import (
    Flask,
    render_template,
    redirect,
    request,
    jsonify,
    send_file,
    session,
    url_for,
)

from kiteconnect import KiteConnect, KiteTicker

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment

import psycopg
from psycopg.types.json import Jsonb


# ============================================================
# CONFIGURATION
# ============================================================

APP_SECRET = os.environ.get("APP_SECRET", "change-me")
KITE_API_KEY = os.environ.get("KITE_API_KEY", "")
KITE_API_SECRET = os.environ.get("KITE_API_SECRET", "")
PUBLIC_BASE_URL = os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")
DATABASE_URL = os.environ.get("DATABASE_URL", "")
PORTAL_USERNAME = os.environ.get("PORTAL_USERNAME", "")
PORTAL_PASSWORD = os.environ.get("PORTAL_PASSWORD", "")

app = Flask(__name__)
app.secret_key = APP_SECRET
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=True,
    PERMANENT_SESSION_LIFETIME=timedelta(days=7),
)

BASE = Path(__file__).resolve().parent
DATA = BASE / "data"
DATA.mkdir(exist_ok=True)

ACCESS_TOKEN_FILE = DATA / "access_token.json"
BASELINE_FILE = DATA / "oi_baseline.json"

HISTORY_DIR = DATA / "history"
HISTORY_DIR.mkdir(exist_ok=True)

lock = threading.RLock()

IST_OFFSET = timedelta(hours=5, minutes=30)

MARKET_START_HOUR = 9
MARKET_START_MINUTE = 15
MARKET_END_HOUR = 15
MARKET_END_MINUTE = 30


# ============================================================
# STATE
# ============================================================

state = {
    "configured": bool(KITE_API_KEY and KITE_API_SECRET),
    "connected": False,
    "message": "Waiting for Zerodha login",
    "last_update": None,

    "date": None,
    "expiry": None,

    "nifty": {
        "price": None,
        "previous_close": None,
        "open": None,
        "high": None,
        "low": None,
        "change": None,
        "change_pct": None,
    },

    "vix": {
        "price": None,
        "range": None,
        "interpretation": None,
    },

    "zone": {
        "quadrant": None,
        "zone": None,
        "opening_pct": None,
        "levels": {},
    },

    "opening_atm": None,

    "oic": {
        "atm": None,
        "minus100": None,
        "plus100": None,
    },

    "series": {
        "nifty": [],
        "atm": [],
        "minus100": [],
        "plus100": [],
        "cio": [],
    },

    "history_dates": [],

    # Strategy trade ledger. The live S1/S2/PNA engine will populate
    # these lists once the exact signal rules are activated.
    "strategies": {
        "S1": {"trades": []},
        "S2": {"trades": []},
        "PNA": {"trades": []},
    },
}


# ============================================================
# GLOBAL LIVE OBJECTS
# ============================================================

kite = None
ticker = None

nifty_token = None
vix_token = None

option_instruments = []
token_meta = {}
oic_tokens = {}

latest_oi = {}
prev_oi = {}

baseline_ready = False
baseline_thread_started = False
snapshot_thread_started = False

latest_nifty = {}
latest_vix = {}

tick_counter = 0
oi_tick_counter = 0


# ============================================================
# TIME
# ============================================================

def now_ist():
    return datetime.utcnow() + IST_OFFSET


def today_key():
    return now_ist().strftime("%Y-%m-%d")


def minute_label():
    return now_ist().strftime("%H:%M")


def is_market_session():
    n = now_ist()

    current = n.hour * 60 + n.minute
    start = MARKET_START_HOUR * 60 + MARKET_START_MINUTE
    end = MARKET_END_HOUR * 60 + MARKET_END_MINUTE

    return start <= current <= end


# ============================================================
# JSON
# ============================================================

def safe_json_load(path, default=None):
    if default is None:
        default = {}

    try:
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)

    except Exception as e:
        print(f"[DIAG] JSON load error: {e}", flush=True)

    return default


def safe_json_save(path, data):
    try:
        temp = path.with_suffix(path.suffix + ".tmp")

        with open(temp, "w", encoding="utf-8") as f:
            json.dump(
                data,
                f,
                ensure_ascii=False,
                indent=2,
            )

        temp.replace(path)

    except Exception as e:
        print(f"[DIAG] JSON save error: {e}", flush=True)


# ============================================================
# PERMANENT DATABASE (NEON POSTGRESQL)
# ============================================================

def db_enabled():
    return bool(DATABASE_URL)


def db_connect():
    if not DATABASE_URL:
        return None

    return psycopg.connect(
        DATABASE_URL,
        connect_timeout=10,
    )


def init_db():
    if not db_enabled():
        print(
            "[DB] DATABASE_URL not configured; using local history fallback.",
            flush=True,
        )
        return False

    try:
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS pratik_daily_history (
                        day DATE PRIMARY KEY,
                        payload JSONB NOT NULL,
                        updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS pratik_kite_session (
                        session_id SMALLINT PRIMARY KEY CHECK (session_id = 1),
                        token_day DATE NOT NULL,
                        access_token TEXT NOT NULL,
                        updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
            conn.commit()

        print(
            "[DB] Neon history table ready.",
            flush=True,
        )
        return True

    except Exception as e:
        print(
            f"[DB] Database initialization failed: {e}",
            flush=True,
        )
        return False


def save_history_to_db(day, payload):
    if not db_enabled():
        return False

    try:
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO pratik_daily_history (day, payload, updated_at)
                    VALUES (%s, %s, NOW())
                    ON CONFLICT (day)
                    DO UPDATE SET
                        payload = EXCLUDED.payload,
                        updated_at = NOW()
                    """,
                    (day, Jsonb(payload)),
                )
            conn.commit()

        return True

    except Exception as e:
        print(
            f"[DB] History save failed for {day}: {e}",
            flush=True,
        )
        return False


def load_history_from_db(day):
    if not db_enabled():
        return None

    try:
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT payload
                    FROM pratik_daily_history
                    WHERE day = %s
                    """,
                    (day,),
                )
                row = cur.fetchone()

        if not row:
            return None

        payload = row[0]

        if isinstance(payload, str):
            payload = json.loads(payload)

        return payload

    except Exception as e:
        print(
            f"[DB] History load failed for {day}: {e}",
            flush=True,
        )
        return None


def db_history_dates():
    if not db_enabled():
        return []

    try:
        with db_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT day
                    FROM pratik_daily_history
                    ORDER BY day DESC
                    """
                )
                rows = cur.fetchall()

        return [row[0].isoformat() for row in rows]

    except Exception as e:
        print(
            f"[DB] History date load failed: {e}",
            flush=True,
        )
        return []


def history_exists(day):
    if load_history_from_db(day) is not None:
        return True

    return history_file(day).exists()


# ============================================================
# HISTORY
# ============================================================

def history_file(day):
    return HISTORY_DIR / f"{day}.json"


def load_day_history(day):
    db_data = load_history_from_db(day)

    if db_data is not None:
        return db_data

    return safe_json_load(
        history_file(day),
        {
            "date": day,
            "expiry": None,
            "opening_atm": None,
            "nifty": {},
            "vix": {},
            "zone": {},
            "oic": {},
            "series": {
                "nifty": [],
                "atm": [],
                "minus100": [],
                "plus100": [],
                "cio": [],
            },
            "strategies": {
                "S1": {"trades": []},
                "S2": {"trades": []},
                "PNA": {"trades": []},
            },
        },
    )


def refresh_history_dates():
    dates = set(db_history_dates())

    for f in HISTORY_DIR.glob("*.json"):
        dates.add(f.stem)

    dates = sorted(
        dates,
        reverse=True,
    )

    with lock:
        state["history_dates"] = dates


def save_current_history():
    with lock:
        day = state.get("date") or today_key()

        payload = {
            "date": day,
            "expiry": state.get("expiry"),
            "opening_atm": state.get("opening_atm"),
            "nifty": state.get("nifty", {}),
            "vix": state.get("vix", {}),
            "zone": state.get("zone", {}),
            "oic": state.get("oic", {}),
            "series": state.get("series", {}),
            "strategies": state.get("strategies", {}),
            "saved_at": now_ist().isoformat(),
        }

    # Local file remains as a temporary fallback.
    safe_json_save(
        history_file(day),
        payload,
    )

    # Neon PostgreSQL is the permanent source of truth.
    save_history_to_db(
        day,
        payload,
    )

    refresh_history_dates()


# ============================================================
# ACCESS TOKEN
# ============================================================
# The daily Kite access token is persisted in Neon so a Render restart
# does not depend on Render's temporary local filesystem.  We keep the
# local JSON file only as a fallback when the database is unavailable.

def save_access_token(token):
    token_day = now_ist().date()
    saved_to_db = False

    if db_enabled():
        try:
            with db_connect() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO pratik_kite_session
                            (session_id, token_day, access_token, updated_at)
                        VALUES (1, %s, %s, NOW())
                        ON CONFLICT (session_id)
                        DO UPDATE SET
                            token_day = EXCLUDED.token_day,
                            access_token = EXCLUDED.access_token,
                            updated_at = NOW()
                        """,
                        (token_day, token),
                    )
                conn.commit()
            saved_to_db = True
            print("[KITE] Today's access token saved to Neon.", flush=True)
        except Exception as e:
            print(f"[KITE] Neon token save failed: {e}", flush=True)

    # Fallback copy. It is not relied on across Render restarts.
    safe_json_save(
        ACCESS_TOKEN_FILE,
        {
            "access_token": token,
            "token_day": token_day.isoformat(),
            "saved_at": now_ist().isoformat(),
            "database_saved": saved_to_db,
        },
    )


def load_access_token():
    today = now_ist().date()

    if db_enabled():
        try:
            with db_connect() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        SELECT access_token
                        FROM pratik_kite_session
                        WHERE session_id = 1 AND token_day = %s
                        """,
                        (today,),
                    )
                    row = cur.fetchone()
            if row and row[0]:
                print("[KITE] Restored today's access token from Neon.", flush=True)
                return row[0]
        except Exception as e:
            print(f"[KITE] Neon token load failed: {e}", flush=True)

    # Local fallback is accepted only if it belongs to today.
    data = safe_json_load(ACCESS_TOKEN_FILE, {})
    token = data.get("access_token")
    token_day = data.get("token_day")
    if token and token_day == today.isoformat():
        print("[KITE] Restored today's access token from local fallback.", flush=True)
        return token

    return None


def clear_access_token():
    if db_enabled():
        try:
            with db_connect() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "DELETE FROM pratik_kite_session WHERE session_id = 1"
                    )
                conn.commit()
        except Exception as e:
            print(f"[KITE] Neon token clear failed: {e}", flush=True)

    try:
        ACCESS_TOKEN_FILE.unlink(missing_ok=True)
    except Exception:
        pass


def restore_kite_session():
    """Reconnect server-side after a Render restart, without a browser."""
    if not (KITE_API_KEY and KITE_API_SECRET):
        return False

    token = load_access_token()
    if not token:
        return False

    try:
        start_live(token)
        print("[KITE] Server-side session restored successfully.", flush=True)
        return True
    except Exception as e:
        print(f"[KITE] Saved session restore failed: {e}", flush=True)
        clear_access_token()
        with lock:
            state["connected"] = False
            state["message"] = "Login required"
        return False


# ============================================================
# ZONE ENGINE
# ============================================================

def round_to_100(value):
    return int(
        round(float(value) / 100.0) * 100
    )


def calculate_zone(open_price, previous_close):
    if not open_price or not previous_close:
        return {
            "quadrant": None,
            "zone": None,
            "opening_pct": None,
            "levels": {},
        }

    opening_pct = (
        (open_price - previous_close)
        / previous_close
        * 100
    )

    quadrant = "Q1" if opening_pct >= 0 else "Q2"

    abs_pct = abs(opening_pct)

    if abs_pct <= 0.25:
        zone = "Z0"
    elif abs_pct <= 0.50:
        zone = "Z1"
    elif abs_pct <= 0.75:
        zone = "Z2"
    elif abs_pct <= 1.00:
        zone = "Z3"
    else:
        zone = "Outside Z3"

    percentages = [
        -1.00,
        -0.75,
        -0.50,
        -0.25,
        0.00,
        0.25,
        0.50,
        0.75,
        1.00,
    ]

    levels = {}

    for pct in percentages:
        value = previous_close * (
            1 + pct / 100
        )

        if pct > 0:
            key = f"+{pct:.2f}%"
        else:
            key = f"{pct:.2f}%"

        levels[key] = round(value, 2)

    return {
        "quadrant": quadrant,
        "zone": zone,
        "opening_pct": round(
            opening_pct,
            3,
        ),
        "levels": levels,
    }


# ============================================================
# INDIA VIX
# ============================================================

def classify_vix(value):
    if value is None:
        return None, None

    value = float(value)

    if value < 12:
        return (
            "<12 LOW",
            "Low volatility. Market may be relatively calm; option premiums can be lower.",
        )

    if value < 15:
        return (
            "12–15 NORMAL",
            "Normal volatility zone.",
        )

    if value < 20:
        return (
            "15–20 ELEVATED",
            "Elevated volatility. Expect larger intraday movement.",
        )

    if value < 25:
        return (
            "20–25 HIGH",
            "High volatility. Use additional caution.",
        )

    return (
        "≥25 VERY HIGH",
        "Very high volatility. Large and rapid market movement is possible.",
    )


# ============================================================
# INSTRUMENT DISCOVERY
# ============================================================

def discover_instruments(k):
    global nifty_token
    global vix_token
    global option_instruments
    global token_meta

    print(
        "[DIAG] Loading instruments...",
        flush=True,
    )

    nse = k.instruments("NSE")
    nfo = k.instruments("NFO")

    print(
        f"[DIAG] NSE={len(nse)} NFO={len(nfo)}",
        flush=True,
    )

    for row in nse:
        symbol = str(
            row.get("tradingsymbol", "")
        ).upper()

        name = str(
            row.get("name", "")
        ).upper()

        if symbol == "NIFTY 50" or name == "NIFTY 50":
            nifty_token = int(
                row["instrument_token"]
            )

        if symbol == "INDIA VIX" or name == "INDIA VIX":
            vix_token = int(
                row["instrument_token"]
            )

    today = now_ist().date()

    candidates = []

    for row in nfo:
        name = str(
            row.get("name", "")
        ).upper()

        inst_type = str(
            row.get("instrument_type", "")
        ).upper()

        if name != "NIFTY":
            continue

        if inst_type not in ("CE", "PE"):
            continue

        expiry = row.get("expiry")

        if not expiry:
            continue

        if isinstance(expiry, str):
            try:
                expiry = datetime.strptime(
                    expiry,
                    "%Y-%m-%d",
                ).date()
            except Exception:
                continue

        if expiry < today:
            continue

        candidates.append(row)

    if not candidates:
        raise RuntimeError(
            "No active NIFTY option contracts found."
        )

    nearest_expiry = min(
        row["expiry"]
        for row in candidates
    )

    option_instruments = [
        row
        for row in candidates
        if row["expiry"] == nearest_expiry
    ]

    token_meta = {}

    for row in option_instruments:
        token = int(
            row["instrument_token"]
        )

        token_meta[token] = {
            "strike": int(
                float(row["strike"])
            ),
            "type": row["instrument_type"],
            "symbol": row["tradingsymbol"],
            "expiry": str(row["expiry"]),
        }

    with lock:
        state["expiry"] = str(
            nearest_expiry
        )

    print(
        f"[DIAG] NIFTY token={nifty_token}",
        flush=True,
    )

    print(
        f"[DIAG] VIX token={vix_token}",
        flush=True,
    )

    print(
        f"[DIAG] Nearest expiry={nearest_expiry}",
        flush=True,
    )

    print(
        f"[DIAG] Option contracts loaded={len(option_instruments)}",
        flush=True,
    )


# ============================================================
# OIC STRIKE LOCK
# ============================================================

def lock_oic_strikes():
    global oic_tokens

    with lock:
        atm = state.get(
            "opening_atm"
        )

    if not atm:
        print(
            "[DIAG] Opening ATM missing",
            flush=True,
        )
        return False

    wanted = {
        "atm": atm,
        "minus100": atm - 100,
        "plus100": atm + 100,
    }

    mapping = {}

    for key, strike in wanted.items():
        ce_token = None
        pe_token = None

        for token, meta in token_meta.items():
            if meta["strike"] != strike:
                continue

            if meta["type"] == "CE":
                ce_token = token

            elif meta["type"] == "PE":
                pe_token = token

        if ce_token and pe_token:
            mapping[key] = {
                "strike": strike,
                "CE": ce_token,
                "PE": pe_token,
            }

    oic_tokens = mapping

    with lock:
        state["oic"]["atm"] = atm
        state["oic"]["minus100"] = atm - 100
        state["oic"]["plus100"] = atm + 100

    print(
        f"[DIAG] OIC strike mapping={mapping}",
        flush=True,
    )

    return len(mapping) == 3


# ============================================================
# PREVIOUS DAY OI BASELINE
# ============================================================

def previous_oi_for_token(k, token):
    end = (
        now_ist().date()
        - timedelta(days=1)
    )

    start = (
        end
        - timedelta(days=10)
    )

    try:
        candles = k.historical_data(
            token,
            start,
            end,
            "day",
            oi=True,
        )

        if not candles:
            return None

        for candle in reversed(candles):
            if candle.get("oi") is not None:
                return int(candle["oi"])

    except Exception as e:
        print(
            f"[DIAG] Historical OI failed token={token}: {e}",
            flush=True,
        )

    return None


def build_oi_baseline():
    global baseline_ready
    global baseline_thread_started
    global prev_oi

    baseline_thread_started = True

    try:
        cached = safe_json_load(
            BASELINE_FILE,
            {},
        )

        cache_date = cached.get("date")
        cache_expiry = cached.get("expiry")

        with lock:
            current_expiry = state.get(
                "expiry"
            )

        if (
            cache_date == today_key()
            and cache_expiry == current_expiry
            and cached.get("oi")
        ):
            prev_oi = {
                int(k): int(v)
                for k, v in cached[
                    "oi"
                ].items()
            }

            baseline_ready = True

            print(
                f"[DIAG] CIO baseline loaded. Contracts={len(prev_oi)}",
                flush=True,
            )

            return

        result = {}

        for index, row in enumerate(
            option_instruments,
            start=1,
        ):
            token = int(
                row["instrument_token"]
            )

            value = previous_oi_for_token(
                kite,
                token,
            )

            if value is not None:
                result[token] = value

            if index % 20 == 0:
                print(
                    f"[DIAG] Baseline {index}/{len(option_instruments)}",
                    flush=True,
                )

            time.sleep(0.35)

        prev_oi = result

        safe_json_save(
            BASELINE_FILE,
            {
                "date": today_key(),
                "expiry": state.get(
                    "expiry"
                ),
                "oi": {
                    str(k): v
                    for k, v in result.items()
                },
            },
        )

        baseline_ready = True

        print(
            f"[DIAG] CIO baseline ready. Contracts={len(prev_oi)}",
            flush=True,
        )

    except Exception as e:
        baseline_ready = False

        print(
            f"[DIAG] CIO baseline ERROR: {e}",
            flush=True,
        )

    finally:
        baseline_thread_started = False


# ============================================================
# CIO
# ============================================================

def cio_totals():
    ce = 0
    pe = 0

    for token, current in latest_oi.items():
        baseline = prev_oi.get(token)
        meta = token_meta.get(token)

        if baseline is None or not meta:
            continue

        delta = int(current) - int(
            baseline
        )

        if delta >= 0:
            continue

        if meta["type"] == "CE":
            ce += delta

        elif meta["type"] == "PE":
            pe += delta

    return ce, pe


# ============================================================
# SERIES
# ============================================================

def append_or_replace_minute(
    series,
    point,
):
    if not series:
        series.append(point)
        return

    if (
        series[-1].get("time")
        == point.get("time")
    ):
        series[-1] = point

    else:
        series.append(point)


def oic_point(key):
    legs = oic_tokens.get(key)

    if not legs:
        return None

    ce = latest_oi.get(
        legs["CE"]
    )

    pe = latest_oi.get(
        legs["PE"]
    )

    if ce is None or pe is None:
        return None

    return {
        "time": minute_label(),
        "timestamp": now_ist().isoformat(),
        "ce": int(ce),
        "pe": int(pe),
    }


def make_snapshot():
    if not state.get("connected"):
        return

    with lock:
        nifty_price = state.get("nifty", {}).get("price")
        if nifty_price is not None:
            nifty_point = {
                "time": minute_label(),
                "timestamp": now_ist().isoformat(),
                "price": float(nifty_price),
            }
            append_or_replace_minute(
                state["series"]["nifty"],
                nifty_point,
            )

        for key in (
            "atm",
            "minus100",
            "plus100",
        ):
            point = oic_point(key)

            if point:
                append_or_replace_minute(
                    state["series"][key],
                    point,
                )

        if baseline_ready:
            ce, pe = cio_totals()

            point = {
                "time": minute_label(),
                "timestamp": now_ist().isoformat(),
                "ce": int(ce),
                "pe": int(pe),
            }

            append_or_replace_minute(
                state["series"]["cio"],
                point,
            )

        state["last_update"] = (
            now_ist().isoformat()
        )

    save_current_history()


def snapshot_worker():
    print(
        "[DIAG] Snapshot worker started",
        flush=True,
    )

    while True:
        try:
            if (
                state.get("connected")
                and is_market_session()
            ):
                make_snapshot()

        except Exception as e:
            print(
                f"[DIAG] Snapshot ERROR: {e}",
                flush=True,
            )

        n = now_ist()

        seconds_to_next = (
            60 - n.second
        )

        if seconds_to_next < 2:
            seconds_to_next = 2

        time.sleep(seconds_to_next)


def ensure_snapshot_worker():
    global snapshot_thread_started

    if snapshot_thread_started:
        return

    snapshot_thread_started = True

    threading.Thread(
        target=snapshot_worker,
        daemon=True,
    ).start()


# ============================================================
# WEBSOCKET
# ============================================================

def on_ticks(ws, ticks):
    global latest_nifty
    global latest_vix
    global tick_counter
    global oi_tick_counter

    tick_counter += len(ticks)

    option_oi_in_batch = 0

    for q in ticks:
        token = int(
            q.get(
                "instrument_token",
                0,
            )
        )

        # NIFTY
        if token == nifty_token:
            price = q.get(
                "last_price"
            )

            ohlc = q.get(
                "ohlc"
            ) or {}

            open_price = ohlc.get(
                "open"
            )

            high = ohlc.get(
                "high"
            )

            low = ohlc.get(
                "low"
            )

            previous_close = ohlc.get(
                "close"
            )

            if price is not None:
                latest_nifty[
                    "price"
                ] = float(price)

            if open_price is not None:
                latest_nifty[
                    "open"
                ] = float(open_price)

            if high is not None:
                latest_nifty[
                    "high"
                ] = float(high)

            if low is not None:
                latest_nifty[
                    "low"
                ] = float(low)

            if previous_close is not None:
                latest_nifty[
                    "previous_close"
                ] = float(
                    previous_close
                )

            p = latest_nifty.get(
                "price"
            )

            pc = latest_nifty.get(
                "previous_close"
            )

            op = latest_nifty.get(
                "open"
            )

            if p is not None and pc:
                latest_nifty[
                    "change"
                ] = round(
                    p - pc,
                    2,
                )

                latest_nifty[
                    "change_pct"
                ] = round(
                    ((p - pc) / pc) * 100,
                    3,
                )

            with lock:
                state["nifty"] = dict(
                    latest_nifty
                )

            if (
                state.get("opening_atm")
                is None
                and op is not None
            ):
                atm = round_to_100(op)

                with lock:
                    state["opening_atm"] = atm

                    state[
                        "zone"
                    ] = calculate_zone(
                        op,
                        pc,
                    )

                lock_oic_strikes()

            elif op is not None and pc:
                with lock:
                    state[
                        "zone"
                    ] = calculate_zone(
                        op,
                        pc,
                    )

        # VIX
        elif token == vix_token:
            price = q.get(
                "last_price"
            )

            if price is not None:
                value = float(price)

                band, interpretation = (
                    classify_vix(value)
                )

                latest_vix = {
                    "price": value,
                    "range": band,
                    "interpretation": interpretation,
                }

                with lock:
                    state["vix"] = dict(
                        latest_vix
                    )

        # OPTION OI
        if token in token_meta:
            oi = q.get("oi")

            if oi is not None:
                latest_oi[
                    token
                ] = int(oi)

                oi_tick_counter += 1
                option_oi_in_batch += 1

    if option_oi_in_batch:
        print(
            f"[DIAG] OI batch={option_oi_in_batch}, unique={len(latest_oi)}",
            flush=True,
        )

    with lock:
        state["last_update"] = (
            now_ist().isoformat()
        )


def on_connect(ws, response):
    tokens = []

    if nifty_token:
        tokens.append(nifty_token)

    if vix_token:
        tokens.append(vix_token)

    tokens.extend(
        token_meta.keys()
    )

    tokens = list(set(tokens))

    print(
        f"[DIAG] WebSocket connected. Subscribing {len(tokens)} tokens.",
        flush=True,
    )

    if tokens:
        ws.subscribe(tokens)

        ws.set_mode(
            ws.MODE_FULL,
            tokens,
        )

    with lock:
        state["connected"] = True
        state["message"] = (
            "LIVE — Zerodha connected"
        )


def on_close(ws, code, reason):
    print(
        f"[DIAG] WebSocket closed {code} {reason}",
        flush=True,
    )

    with lock:
        state["connected"] = False
        state["message"] = (
            f"Disconnected — {reason or code}"
        )


def on_error(ws, code, reason):
    print(
        f"[DIAG] WebSocket ERROR {code} {reason}",
        flush=True,
    )

    with lock:
        state["message"] = (
            f"WebSocket error ({code}) — {reason}"
        )


# ============================================================
# START LIVE
# ============================================================

def start_live(access_token):
    global kite
    global ticker
    global baseline_ready

    print(
        "[DIAG] start_live() called",
        flush=True,
    )

    with lock:
        state["date"] = today_key()
        state["message"] = (
            "Starting Zerodha live feed..."
        )

    kite = KiteConnect(
        api_key=KITE_API_KEY
    )

    kite.set_access_token(
        access_token
    )

    kite.profile()

    discover_instruments(kite)

    # Important:
    # if opening ATM was restored from today's
    # history, rebuild OIC token mapping.
    if state.get("opening_atm"):
        lock_oic_strikes()

    baseline_ready = False

    if not baseline_thread_started:
        threading.Thread(
            target=build_oi_baseline,
            daemon=True,
        ).start()

    ticker = KiteTicker(
        KITE_API_KEY,
        access_token,
    )

    ticker.on_ticks = on_ticks
    ticker.on_connect = on_connect
    ticker.on_close = on_close
    ticker.on_error = on_error

    ticker.connect(
        threaded=True
    )

    ensure_snapshot_worker()


# ============================================================
# RESTORE TODAY
# ============================================================

def restore_today_history():
    day = today_key()

    saved = load_day_history(day)

    series = saved.get(
        "series"
    ) or {}

    with lock:
        state["date"] = day

        for key in (
            "nifty",
            "atm",
            "minus100",
            "plus100",
            "cio",
        ):
            if isinstance(
                series.get(key),
                list,
            ):
                state[
                    "series"
                ][key] = series[key]

        if saved.get(
            "opening_atm"
        ):
            state[
                "opening_atm"
            ] = saved[
                "opening_atm"
            ]

        if saved.get("expiry"):
            state["expiry"] = saved[
                "expiry"
            ]

        if saved.get("zone"):
            state["zone"] = saved[
                "zone"
            ]

        if saved.get("nifty"):
            state["nifty"] = saved[
                "nifty"
            ]

        if saved.get("vix"):
            state["vix"] = saved[
                "vix"
            ]

        saved_strategies = saved.get("strategies") or {}
        for strategy_name in ("S1", "S2", "PNA"):
            strategy_data = saved_strategies.get(strategy_name) or {}
            trades = strategy_data.get("trades") or []
            if isinstance(trades, list):
                state["strategies"][strategy_name]["trades"] = trades


init_db()
restore_today_history()
refresh_history_dates()
# Important for free Render: when GitHub Actions wakes/restarts the service,
# reconnect to Kite from today's Neon-persisted token without requiring the
# dashboard to be opened in a browser.
restore_kite_session()


# ============================================================
# ROUTES
# ============================================================

# ============================================================
# PORTAL AUTHENTICATION
# ============================================================

def portal_login_configured():
    return bool(PORTAL_USERNAME and PORTAL_PASSWORD)


def portal_authenticated():
    return bool(session.get("portal_authenticated"))


@app.before_request
def require_portal_login():
    # Login page, static assets and health check stay public.
    if request.endpoint in {"portal_login", "static", "health"}:
        return None

    if not portal_authenticated():
        next_url = request.full_path if request.query_string else request.path
        return redirect(url_for("portal_login", next=next_url))

    return None


@app.route("/login", methods=["GET", "POST"])
def portal_login():
    if portal_authenticated():
        return redirect(url_for("index"))

    error = None
    configured = portal_login_configured()

    if request.method == "POST":
        if not configured:
            error = "Portal login is not configured on the server."
        else:
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")

            user_ok = hmac.compare_digest(username, PORTAL_USERNAME)
            pass_ok = hmac.compare_digest(password, PORTAL_PASSWORD)

            if user_ok and pass_ok:
                session.clear()
                session["portal_authenticated"] = True
                session["portal_username"] = PORTAL_USERNAME
                session.permanent = request.form.get("remember") == "on"

                next_url = request.args.get("next", "")
                if not next_url.startswith("/") or next_url.startswith("//"):
                    next_url = url_for("index")
                return redirect(next_url)

            error = "Invalid username or password."

    return render_template(
        "login.html",
        error=error,
        configured=configured,
    )


@app.route("/logout")
def portal_logout():
    session.clear()
    return redirect(url_for("portal_login"))


@app.route("/")
def index():
    if (
        not KITE_API_KEY
        or not KITE_API_SECRET
    ):
        return render_template(
            "index.html",
            configured=False,
            base_url=PUBLIC_BASE_URL,
        )

    token = load_access_token()

    if token and not state["connected"]:
        try:
            start_live(token)

        except Exception as e:
            print(
                f"[DIAG] Existing token failed: {e}",
                flush=True,
            )

            clear_access_token()

            with lock:
                state["connected"] = False
                state["message"] = (
                    "Login required"
                )

    return render_template(
        "index.html",
        configured=True,
        base_url=PUBLIC_BASE_URL,
    )


@app.route("/kite/login")
def kite_login():
    if not KITE_API_KEY:
        return (
            "KITE_API_KEY is not configured",
            400,
        )

    k = KiteConnect(
        api_key=KITE_API_KEY
    )

    return redirect(
        k.login_url()
    )


@app.route("/kite/callback")
def kite_callback():
    request_token = request.args.get(
        "request_token"
    )

    if not request_token:
        return (
            "Zerodha did not return a request_token",
            400,
        )

    k = KiteConnect(
        api_key=KITE_API_KEY
    )

    session = k.generate_session(
        request_token,
        api_secret=KITE_API_SECRET,
    )

    access_token = session[
        "access_token"
    ]

    save_access_token(
        access_token
    )

    start_live(
        access_token
    )

    return redirect("/")


@app.route("/kite/logout")
def kite_logout():
    global ticker

    try:
        if ticker:
            ticker.close()

    except Exception as e:
        print(
            f"[DIAG] Ticker close error: {e}",
            flush=True,
        )

    clear_access_token()

    with lock:
        state["connected"] = False
        state["message"] = (
            "Logged out — Zerodha login required"
        )

    return redirect("/")


@app.route("/api/state")
def api_state():
    with lock:
        return jsonify(state)


@app.route("/api/history/dates")
def api_history_dates():
    refresh_history_dates()

    with lock:
        return jsonify(
            {
                "dates": state[
                    "history_dates"
                ]
            }
        )


@app.route("/api/history/<day>")
def api_history_day(day):
    if not history_exists(day):
        return jsonify(
            {
                "error": (
                    "No stored data for selected date."
                )
            }
        ), 404

    return jsonify(
        load_day_history(day)
    )


@app.route("/api/history/<day>/cio")
def api_history_cio(day):
    # For today use current live memory.
    if day == today_key():
        with lock:
            return jsonify(
                {
                    "date": day,
                    "expiry": state.get(
                        "expiry"
                    ),
                    "opening_atm": state.get(
                        "opening_atm"
                    ),
                    "cio": list(
                        state.get(
                            "series",
                            {},
                        ).get(
                            "cio",
                            [],
                        )
                    ),
                }
            )

    if not history_exists(day):
        return jsonify(
            {
                "error": (
                    "No stored data for selected date."
                )
            }
        ), 404

    data = load_day_history(day)

    return jsonify(
        {
            "date": day,
            "expiry": data.get(
                "expiry"
            ),
            "opening_atm": data.get(
                "opening_atm"
            ),
            "cio": data.get(
                "series",
                {},
            ).get(
                "cio",
                [],
            ),
        }
    )


# ============================================================
# OIC + CIO + NIFTY EXCEL DOWNLOAD
# Existing route is preserved for frontend compatibility.
# ============================================================

@app.route("/api/download/cio/<day>")
def download_cio_excel(day):
    if day == today_key():
        with lock:
            data = {
                "date": day,
                "expiry": state.get("expiry"),
                "opening_atm": state.get("opening_atm"),
                "series": {
                    key: list(state.get("series", {}).get(key, []))
                    for key in ("nifty", "minus100", "atm", "plus100", "cio")
                },
            }
    else:
        if not history_exists(day):
            return jsonify({"error": "No stored data for selected date."}), 404
        data = load_day_history(day)

    series = data.get("series", {}) or {}
    nifty_data = series.get("nifty", []) or []
    minus100 = series.get("minus100", []) or []
    atm = series.get("atm", []) or []
    plus100 = series.get("plus100", []) or []
    cio = series.get("cio", []) or []

    if not any((nifty_data, minus100, atm, plus100, cio)):
        return jsonify({"error": "No OIC/CIO/NIFTY data available for selected date."}), 404

    def by_time(rows):
        return {str(p.get("time")): p for p in rows if p.get("time")}

    maps = {
        "nifty": by_time(nifty_data),
        "minus100": by_time(minus100),
        "atm": by_time(atm),
        "plus100": by_time(plus100),
        "cio": by_time(cio),
    }
    all_times = sorted(set().union(*(m.keys() for m in maps.values())))

    wb = Workbook()
    ws = wb.active
    ws.title = "OIC + CIO + NIFTY"

    ws["A1"] = "Pratik Analysis"
    ws["A2"] = "NIFTY + OIC + CIO — Minute-wise Data"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"].font = Font(bold=True, size=13)

    ws["A4"] = "Date"
    ws["B4"] = day
    ws["A5"] = "NIFTY Opening ATM"
    ws["B5"] = data.get("opening_atm")
    ws["A6"] = "Nearest Expiry"
    ws["B6"] = data.get("expiry")

    headers = [
        "Time",
        "NIFTY Spot",
        "ATM -100 CE OI",
        "ATM -100 PE OI",
        "ATM CE OI",
        "ATM PE OI",
        "ATM +100 CE OI",
        "ATM +100 PE OI",
        "CIO CE Negative Change in OI",
        "CIO PE Negative Change in OI",
    ]
    header_row = 8
    for col, value in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row_no, t in enumerate(all_times, start=header_row + 1):
        n = maps["nifty"].get(t, {})
        m = maps["minus100"].get(t, {})
        a = maps["atm"].get(t, {})
        p = maps["plus100"].get(t, {})
        c = maps["cio"].get(t, {})
        values = [
            t, n.get("price"),
            m.get("ce"), m.get("pe"),
            a.get("ce"), a.get("pe"),
            p.get("ce"), p.get("pe"),
            c.get("ce"), c.get("pe"),
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=row_no, column=col, value=value)

    ws.freeze_panes = "A9"
    widths = [14, 16, 20, 20, 20, 20, 20, 20, 32, 32]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Pratik_Analysis_NIFTY_OIC_CIO_{day}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# STRATEGY REPORTS + EXCEL EXPORT
# ============================================================

def _parse_report_day(value):
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except Exception:
        return None


def _strategy_trades_from_day(day, data):
    strategies = (data or {}).get("strategies") or {}
    rows = []

    for strategy_name in ("S1", "S2", "PNA"):
        strategy_data = strategies.get(strategy_name) or {}
        trades = strategy_data.get("trades") or []

        if not isinstance(trades, list):
            continue

        for i, trade in enumerate(trades, start=1):
            if not isinstance(trade, dict):
                continue

            row = dict(trade)
            row.setdefault("date", day)
            row.setdefault("strategy", strategy_name)
            row.setdefault("trade_no", i)
            rows.append(row)

    return rows


def _report_days(start_day, end_day):
    refresh_history_dates()
    with lock:
        available = list(state.get("history_dates", []))
        today_has_data = bool(state.get("date") == today_key())

    if today_has_data and today_key() not in available:
        available.append(today_key())

    selected = []
    for day in available:
        d = _parse_report_day(day)
        if d and start_day <= d <= end_day:
            selected.append(day)

    return sorted(set(selected))


def _collect_report(start_text, end_text):
    start_day = _parse_report_day(start_text)
    end_day = _parse_report_day(end_text)

    if not start_day or not end_day:
        return None, "Please select a valid From and To date."

    if start_day > end_day:
        return None, "From date cannot be after To date."

    days = _report_days(start_day, end_day)
    all_trades = []

    for day in days:
        if day == today_key():
            with lock:
                data = {
                    "date": day,
                    "strategies": json.loads(json.dumps(state.get("strategies", {}))),
                }
        else:
            data = load_day_history(day)

        all_trades.extend(_strategy_trades_from_day(day, data))

    def num(value):
        try:
            return float(value)
        except Exception:
            return 0.0

    summary = {}
    for name in ("S1", "S2", "PNA"):
        trades = [t for t in all_trades if str(t.get("strategy", "")).upper() == name]
        completed = [t for t in trades if t.get("exit_time") or t.get("exit_level") is not None]
        points = [num(t.get("points")) for t in completed]
        wins = sum(1 for p in points if p > 0)
        losses = sum(1 for p in points if p < 0)
        flat = sum(1 for p in points if p == 0)
        sl_hits = sum(1 for t in completed if bool(t.get("sl_hit")))
        maes = [num(t.get("mae")) for t in completed if t.get("mae") is not None]
        mfes = [num(t.get("mfe")) for t in completed if t.get("mfe") is not None]

        summary[name] = {
            "trades": len(completed),
            "wins": wins,
            "losses": losses,
            "flat": flat,
            "win_rate": round((wins / len(completed) * 100), 2) if completed else 0.0,
            "total_points": round(sum(points), 2),
            "avg_points": round((sum(points) / len(completed)), 2) if completed else 0.0,
            "sl_hits": sl_hits,
            "avg_mae": round((sum(maes) / len(maes)), 2) if maes else 0.0,
            "avg_mfe": round((sum(mfes) / len(mfes)), 2) if mfes else 0.0,
            "best_trade": round(max(points), 2) if points else 0.0,
            "worst_trade": round(min(points), 2) if points else 0.0,
        }

    daywise = []
    for day in days:
        item = {"date": day}
        for name in ("S1", "S2", "PNA"):
            day_trades = [
                t for t in all_trades
                if t.get("date") == day
                and str(t.get("strategy", "")).upper() == name
                and (t.get("exit_time") or t.get("exit_level") is not None)
            ]
            item[f"{name}_trades"] = len(day_trades)
            item[f"{name}_points"] = round(sum(num(t.get("points")) for t in day_trades), 2)
        daywise.append(item)

    return {
        "from": start_text,
        "to": end_text,
        "days": days,
        "summary": summary,
        "daywise": daywise,
        "trades": all_trades,
    }, None


@app.route("/api/reports")
def api_strategy_reports():
    start_text = request.args.get("from", "")
    end_text = request.args.get("to", "")
    report, error = _collect_report(start_text, end_text)
    if error:
        return jsonify({"error": error}), 400
    return jsonify(report)


@app.route("/api/download/reports")
def download_strategy_reports_excel():
    start_text = request.args.get("from", "")
    end_text = request.args.get("to", "")
    report, error = _collect_report(start_text, end_text)
    if error:
        return jsonify({"error": error}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Overall Summary"

    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True)

    ws["A1"] = "Pratik Analysis"
    ws["A2"] = "S1 / S2 / PNA Strategy Performance Report"
    ws["A1"].font = title_font
    ws["A2"].font = Font(bold=True, size=13)
    ws["A4"] = "From"
    ws["B4"] = start_text
    ws["C4"] = "To"
    ws["D4"] = end_text

    summary_headers = [
        "Strategy", "Trades", "Wins", "Losses", "Flat", "Win %",
        "Total Points", "Avg Points/Trade", "30-Point SL Hits",
        "Avg MAE", "Avg MFE", "Best Trade", "Worst Trade",
    ]
    for col, value in enumerate(summary_headers, 1):
        cell = ws.cell(row=6, column=col, value=value)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_no, name in enumerate(("S1", "S2", "PNA"), start=7):
        m = report["summary"][name]
        values = [
            name, m["trades"], m["wins"], m["losses"], m["flat"], m["win_rate"],
            m["total_points"], m["avg_points"], m["sl_hits"], m["avg_mae"],
            m["avg_mfe"], m["best_trade"], m["worst_trade"],
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=row_no, column=col, value=value)

    ws_day = wb.create_sheet("Day-wise Summary")
    day_headers = ["Date", "S1 Trades", "S1 Points", "S2 Trades", "S2 Points", "PNA Trades", "PNA Points"]
    for col, value in enumerate(day_headers, 1):
        c = ws_day.cell(row=1, column=col, value=value)
        c.font = header_font
    for row_no, d in enumerate(report["daywise"], start=2):
        vals = [
            d["date"], d["S1_trades"], d["S1_points"], d["S2_trades"],
            d["S2_points"], d["PNA_trades"], d["PNA_points"],
        ]
        for col, value in enumerate(vals, 1):
            ws_day.cell(row=row_no, column=col, value=value)

    trade_headers = [
        "Date", "Trade #", "Type", "Entry Time", "Entry Level", "30-Point SL",
        "Exit Time", "Exit Level", "Points", "MAE", "MFE", "SL Hit?",
        "Result", "Exit Reason", "Comments",
    ]

    for strategy_name in ("S1", "S2", "PNA"):
        sheet = wb.create_sheet(f"{strategy_name} Trades")
        for col, value in enumerate(trade_headers, 1):
            c = sheet.cell(row=1, column=col, value=value)
            c.font = header_font

        rows = [t for t in report["trades"] if str(t.get("strategy", "")).upper() == strategy_name]
        for row_no, t in enumerate(rows, start=2):
            points = t.get("points")
            result = t.get("result")
            if not result and points is not None:
                try:
                    p = float(points)
                    result = "WIN" if p > 0 else "LOSS" if p < 0 else "FLAT"
                except Exception:
                    result = ""

            vals = [
                t.get("date"), t.get("trade_no"), t.get("type") or t.get("side"),
                t.get("entry_time"), t.get("entry_level"), t.get("sl_level"),
                t.get("exit_time"), t.get("exit_level"), points, t.get("mae"),
                t.get("mfe"), "YES" if t.get("sl_hit") else "NO", result,
                t.get("exit_reason"), t.get("comments"),
            ]
            for col, value in enumerate(vals, 1):
                sheet.cell(row=row_no, column=col, value=value)

    risk = wb.create_sheet("Risk & SL")
    risk_headers = ["Strategy", "Trades", "30-Point SL Hits", "Avg MAE", "Avg MFE", "Best Trade", "Worst Trade"]
    for col, value in enumerate(risk_headers, 1):
        c = risk.cell(row=1, column=col, value=value)
        c.font = header_font
    for row_no, name in enumerate(("S1", "S2", "PNA"), start=2):
        m = report["summary"][name]
        vals = [name, m["trades"], m["sl_hits"], m["avg_mae"], m["avg_mfe"], m["best_trade"], m["worst_trade"]]
        for col, value in enumerate(vals, 1):
            risk.cell(row=row_no, column=col, value=value)

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2" if sheet.title != "Overall Summary" else "A6"
        for column_cells in sheet.columns:
            max_len = 0
            letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 11), 30)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Pratik_Analysis_Strategy_Report_{start_text}_to_{end_text}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# HEALTH
# ============================================================

@app.route("/health")
def health():
    return jsonify(
        {
            "ok": True,
            "connected": bool(
                state.get(
                    "connected"
                )
            ),
            "date": state.get(
                "date"
            ),
            "option_tokens": len(
                token_meta
            ),
            "latest_oi_tokens": len(
                latest_oi
            ),
            "baseline_ready": baseline_ready,
            "oic_tokens": oic_tokens,
            "database_configured": db_enabled(),
            "database_history_days": len(db_history_dates()),
            "cio_points": len(
                state.get(
                    "series",
                    {},
                ).get(
                    "cio",
                    [],
                )
            ),
        }
    )


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    port = int(
        os.environ.get(
            "PORT",
            "8000",
        )
    )

    app.run(
        host="0.0.0.0",
        port=port,
        debug=False,
    )
